import json
from io import BytesIO
import openpyxl
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from .models import Item, Quotation   # Quotation is optional – comment out if not used
from django.views.decorators.http import require_POST
from django.utils import timezone

@login_required
def home(request):
    # You can pass any context needed, but splash page likely doesn't need cart
    return render(request, 'home.html')   # home.html extends base without overriding content


# ----- Helper -----
def get_cart_items(request):
    """
    Return a list of dictionaries, each containing an 'item' (Item object)
    and its 'quantity', built from the session cart.
    The session cart must be stored as a list of dicts: [{'item_id': id, 'quantity': qty}, ...]
    """
    cart = request.session.get('cart', [])
    cart_items = []
    for entry in cart:
        try:
            item = Item.objects.get(pk=entry['item_id'])
            cart_items.append({'item': item, 'quantity': entry['quantity']})
        except Item.DoesNotExist:
            continue   # skip invalid entries
    return cart_items

@login_required
def select_items(request):
    cart = request.session.get('cart', [])
    # Build cart items for display
    cart_items = []
    for entry in cart:
        try:
            item = Item.objects.get(pk=entry['item_id'])
            cart_items.append({'item': item, 'quantity': entry['quantity']})
        except Item.DoesNotExist:
            continue
    return render(request, 'select_items.html', {'cart_items': cart_items})


# ----- Authentication Views -----
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ----- Main Views -----



@login_required
def item_detail(request, item_id):
    item = get_object_or_404(Item, pk=item_id)
    cart = request.session.get('cart', [])

    # Add item to cart if not already present (with quantity 1)
    found = False
    for entry in cart:
        if entry['item_id'] == item_id:
            found = True
            break
    if not found:
        cart.append({'item_id': item_id, 'quantity': 1})
        request.session['cart'] = cart
        request.session.modified = True

    cart_items = get_cart_items(request)
    return render(request, 'detail.html', {'item': item, 'cart_items': cart_items})


@login_required
def add_item(request):
    if request.method == 'POST':
        product_name = request.POST.get('product_name')
        category_id = request.POST.get('category')
        company_id = request.POST.get('company')
        try:
            item = Item.objects.get(
                product_name=product_name,
                category_id=category_id,
                company_id=company_id
            )
            return redirect('item_detail', item_id=item.id)
        except Item.DoesNotExist:
            messages.error(request, "Item not found. Please try again.")
            return redirect('home')
    return redirect('home')


# ----- API Endpoints (JSON) -----
@login_required
def api_products(request):
    term = request.GET.get('q', '')
    products = Item.objects.filter(product_name__icontains=term) \
        .values_list('product_name', flat=True).distinct()
    return JsonResponse(list(products), safe=False)


@login_required
def api_categories(request):
    product = request.GET.get('product')
    if not product:
        return JsonResponse([], safe=False)
    categories = Item.objects.filter(product_name=product) \
        .values_list('category__id', 'category__name').distinct()
    data = [{'id': cid, 'name': name} for cid, name in categories]
    return JsonResponse(data, safe=False)


@login_required
def api_companies(request):
    product = request.GET.get('product')
    category_id = request.GET.get('category')
    if not product or not category_id:
        return JsonResponse([], safe=False)
    companies = Item.objects.filter(product_name=product, category_id=category_id) \
        .values_list('company__id', 'company__name').distinct()
    data = [{'id': cid, 'name': name} for cid, name in companies]
    return JsonResponse(data, safe=False)


# ----- Cart Management -----
@login_required
def remove_from_cart(request, item_id):
    if request.method == 'POST':
        cart = request.session.get('cart', [])
        new_cart = [entry for entry in cart if entry['item_id'] != item_id]
        request.session['cart'] = new_cart
        request.session.modified = True
        messages.success(request, "Item removed from cart.")
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required
def update_cart_quantity(request, item_id):
    if request.method == 'POST':
        quantity = request.POST.get('quantity')
        try:
            quantity = int(quantity)
            if quantity < 1:
                quantity = 1
        except (TypeError, ValueError):
            quantity = 1

        cart = request.session.get('cart', [])
        for entry in cart:
            if entry['item_id'] == item_id:
                entry['quantity'] = quantity
                break
        request.session['cart'] = cart
        request.session.modified = True
        messages.success(request, "Quantity updated.")
    return redirect(request.META.get('HTTP_REFERER', 'home'))


# ----- Excel Generation -----
@login_required
def generate_excel(request):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    cart = request.session.get('cart', [])
    if not cart:
        messages.warning(request, "No items in cart to generate Excel.")
        return redirect('home')

    # Fetch items efficiently
    item_ids = [entry['item_id'] for entry in cart]
    items_dict = {
        item.id: item for item in Item.objects.filter(pk__in=item_ids)
                     .select_related('category', 'company')
    }

    wb = Workbook()
    wb.remove(wb.active)

    sheets_config = [
        ("Quality 1 - Premium", "price_q1"),
        ("Quality 2 - Standard", "price_q2"),
        ("Quality 3 - Economy", "price_q3"),
    ]

    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, price_field in sheets_config:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

        headers = [
            "Item",
            "Description",
            "Price (₹)",
            "Quantity",
            "Total (₹)",
            "Labour (₹)",
            "Subtotal (₹)"
        ]

        # Header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")

        row_num = 2
        grand_total = 0.0

        for entry in cart:
            item_id = entry['item_id']
            quantity = entry.get('quantity', 1)
            labour = entry.get('labour', 0.0)  # 0 if not yet saved

            item = items_dict.get(item_id)
            if not item:
                continue

            unit_price = float(getattr(item, price_field, 0.0))
            total = quantity * unit_price
            subtotal = total + labour
            grand_total += subtotal

            description = (
                f"Category: {item.category.name}\n"
                f"Company: {item.company.name}\n"
                f"{item.description or 'No description'}"
            )

            row_data = [
                item.product_name,
                description,
                unit_price,
                quantity,
                total,
                labour,
                subtotal
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center' if col_num >= 3 else 'left',
                    vertical='center',
                    wrap_text=True
                )
                if col_num >= 3:  # numeric columns
                    cell.number_format = '#,##0.00'

            row_num += 1

        # Grand Total
        if row_num > 2:
            ws.cell(row=row_num, column=6, value="Grand Total").font = Font(bold=True)
            grand_cell = ws.cell(row=row_num, column=7, value=grand_total)
            grand_cell.font = Font(bold=True, color="006400")
            grand_cell.number_format = '#,##0.00'
            grand_cell.border = thin_border
            grand_cell.alignment = Alignment(horizontal='center')

        # Auto column width
        for col in range(1, 8):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value or "")) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    # Response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = FileResponse(
        buffer,
        as_attachment=True,
        filename=f"Quotation_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@login_required
@require_POST
def update_labour(request):
    try:
        data = json.loads(request.body)
        item_id = data['item_id']
        labour = float(data['labour'])

        cart = request.session.get('cart', [])
        for entry in cart:
            if str(entry['item_id']) == str(item_id):
                entry['labour'] = labour
                break
        request.session['cart'] = cart
        request.session.modified = True
        return JsonResponse({'status': 'ok'})
    except:
        return JsonResponse({'status': 'error'}, status=400)

# ----- Save Quotation (optional – requires Quotation model) -----
# If you haven't created the Quotation model, comment out the following view and its URL.
@login_required
def save_quotation(request):
    if request.method == 'POST':
        cart = request.session.get('cart', [])
        if not cart:
            messages.warning(request, "No items to save.")
            return redirect('home')

        items_data = []
        for entry in cart:
            try:
                item = Item.objects.get(pk=entry['item_id'])
                items_data.append({
                    'item_id': item.id,
                    'product_name': item.product_name,
                    'category': item.category.name,
                    'company': item.company.name,
                    'quantity': entry['quantity'],
                    'price_q1': float(item.price_q1),
                    'price_q2': float(item.price_q2),
                    'price_q3': float(item.price_q3),
                })
            except Item.DoesNotExist:
                continue

        # Create Quotation object (model must exist)
        quotation = Quotation.objects.create(
            user=request.user,
            items=json.dumps(items_data)
        )
        messages.success(request, "Quotation saved successfully.")
        return redirect('home')
    return redirect('home')